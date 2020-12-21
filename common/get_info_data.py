#_*_encoding=utf-8_*_
#作者     :bozhong
#创建时间 :2020/10/1415:14
#文件     :
#IDE      :PyCharm
from openpyxl import load_workbook
from common import project_path,setting
import random



class GetInfoData:

    RID=None
    TID=None
    CID=None

    def __init__(self,sheet_name='info'):
        self.sheet_name=sheet_name
        self.wb=load_workbook(project_path.test_case_path)
        self.sheet=[sheet_name]
        self.login_phone=self.wb[ self.sheet_name].cell(1,2).value


    def get_community_id(self,flag=1,type='tid',write_id=None):
        '''
        :param flag:是否存取  0 存 1取  默认取
        :param tpye: 类型  tid cid rid
        :param write_id:写入的id
        :return: 对应的id
        '''
        self.sheet = self.wb[self.sheet_name]
        if flag ==0:
            if type == 'cid':
                self.sheet.cell(6,2).value = write_id
            elif type == 'rid':
                self.sheet.cell(7, 2).value = write_id
            else:
                self.sheet.cell(5, 2).value = write_id
            self.wb.save('info')
        else:
            if type == 'cid':
                id=self.wb[self.sheet_name].cell(6,2).value
            elif type == 'rid':
                id = self.wb[self.sheet_name].cell(7, 2).value
            else:
                id = self.wb[self.sheet_name].cell(5, 2).value
            return id

    def get_token(self):
        if setting.ENVIRONMENT == 3:
            token = self.wb['info'].cell(2, 2).value
        elif setting.ENVIRONMENT == 2:
            token = self.wb['info'].cell(4, 2).value
        else:
            token = self.wb['info'].cell(3, 2).value
        return token

    def get_hander(self):
        handers=[{'User-Agent':'bz-mindfulness-android-1.5.1 from android sdk version:24 mobile versionCode:7.0 phone model:PRO 7-H BZChannelNo-baidu',
        'Content-Type':'application/x-www-form-urlencoded'},
                 {'User-Agent':'bz-mindfulness-android-1.5.1 from android sdk version:28 mobile versionCode:9 phone model:MI 8 Lite BZChannelNo-xiaomib'},
                 ]
        return random.choice(handers)

    def BaseUrl(self,num=None):
        if num==3:
            base_url = 'bozhong.com'
        elif num==2:
            base_url = 'online.seedit.cc'
        else:
            base_url = 'office.bzdev.net'
        return base_url



#环境切换
def Application_profiles(data):
    res_data=[]
    for item in data:
        item['url'] = item['url'].replace('office.bzdev.net', GetInfoData().BaseUrl(setting.ENVIRONMENT))
        res_data.append(item)
    return res_data



def replace_data(one_data,two_data=setting.REPALCE_DATA):
    '''
    替换数据，一定要保持GetInfoData里有对应的值
    :param one_data: 原数据
    :param two_data:需要替换的键
    :return:替换后的数据
    '''
    for item in two_data:
        if one_data['param'].find('${%s}'%item)!=-1:
            one_data['param'] = one_data['param'].replace('${%s}'%item, str(getattr(GetInfoData, item.upper())))
    return one_data


if __name__ == '__main__':
    # print(type(GetInfoData.base_url))
    # print(GetInfoData().BaseUrl(setting.ENVIRONMENT))
    # obj=GetInfoData()
    # obj.get_community_id(0,'cid','111181')
    # print(obj.get_community_id(1))

    one_data={'param':"{'id':1,'tid':'${tid}','cid':${cid}}"}

    res=replace_data(one_data)
    print(res)
