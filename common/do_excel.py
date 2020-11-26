#_*_encoding=utf-8_*_
#作者     :bozhong
#创建时间 :2020/4/1014:49
#文件     :
#IDE      :PyCharm
from openpyxl import load_workbook
from common import project_path
from common.get_info_data import GetInfoData,Application_profiles
class DoExce():
    def __init__(self,file_name):
        self.file_name=file_name

    #读取Excel的数据
    def get_excel(self,sheetname):
        wb = load_workbook(self.file_name)
        sheet=wb[sheetname]
        #
        # if sheet_name=='login':
        #     phone=sheet.cell(1, 2).value
        #     setattr(GetInfoData,'login_phone',phone)
        test_data=[]
        for i in range(2,sheet.max_row+1):
            sub_data={}
            sub_data['id'] = sheet.cell(i, 1).value  # 用例序号
            sub_data['HttpMethod'] = sheet.cell(i, 2).value
            sub_data['module'] = sheet.cell(i, 3).value
            sub_data['description'] = sheet.cell(i, 4).value
            sub_data['url'] = sheet.cell(i, 5).value

            if sheet.cell(i,6).value.find('${phone}')!=-1:
                sub_data['param'] = sheet.cell(i, 6).value.replace('${phone}',str(GetInfoData().login_phone))
            else:
                sub_data['param'] = sheet.cell(i,6).value
            sub_data['ExpectedResult']=sheet.cell(i,7).value
            test_data.append(sub_data)
        return test_data

    def write_back(self,sheet_name,row,ActualResult,TestResult,Comment):
        wb=load_workbook(self.file_name)
        sheet=wb[sheet_name]
        sheet.cell(row,8).value=ActualResult
        sheet.cell(row, 9).value = TestResult
        sheet.cell(row, 10).value = Comment

        #保存
        wb.save(self.file_name)

if __name__=='__main__':
    import json
    sheet_name='mindfulness'
    wb=DoExce(project_path.test_case_path).get_excel(sheet_name)
    # data=(wb[0]['param'])
    # print(wb)
    # for i in wb:
        # i['url']=i['url'].replace('office.bzdev.net',GetInfoData.base_url)
        # print(i)
    data=Application_profiles(wb)
    print(data)